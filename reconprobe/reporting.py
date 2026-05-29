"""Reporting Automation module for ReconProbe — Phase 6.

Provides advanced reporting capabilities including:
- CVSS v3.1 base score calculation
- Professional PDF report generation
- CSV/CSV export of findings
- Executive summary generation
- Scan timeline and metrics
"""

from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

try:
    from fpdf import FPDF

    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ────────────────────────────────────────────────────────────────────────────
# CVSS v3.1 Base Score Calculator
# ────────────────────────────────────────────────────────────────────────────

# CVSS v3.1 metric weights
_CVSS_METRICS: dict[str, dict[str, float]] = {
    "attack_vector": {"N": 0.85, "A": 0.62, "L": 0.55, "P": 0.20},
    "attack_complexity": {"L": 0.77, "H": 0.44},
    "privileges_required": {"N": 0.85, "L": 0.62, "H": 0.27},
    "user_interaction": {"N": 0.85, "R": 0.62},
    "scope": {"U": 0.0, "C": 1.0},
    "confidentiality": {"H": 0.56, "L": 0.22, "N": 0.0},
    "integrity": {"H": 0.56, "L": 0.22, "N": 0.0},
    "availability": {"H": 0.56, "L": 0.22, "N": 0.0},
}

_CVSS_LABELS: dict[str, dict[str, str]] = {
    "attack_vector": {
        "N": "Network",
        "A": "Adjacent Network",
        "L": "Local",
        "P": "Physical",
    },
    "attack_complexity": {"L": "Low", "H": "High"},
    "privileges_required": {"N": "None", "L": "Low", "H": "High"},
    "user_interaction": {"N": "None", "R": "Required"},
    "scope": {"U": "Unchanged", "C": "Changed"},
    "confidentiality": {"H": "High", "L": "Low", "N": "None"},
    "integrity": {"H": "High", "L": "Low", "N": "None"},
    "availability": {"H": "High", "L": "Low", "N": "None"},
}


def calculate_cvss_score(
    attack_vector: str = "N",
    attack_complexity: str = "L",
    privileges_required: str = "N",
    user_interaction: str = "N",
    scope: str = "U",
    confidentiality: str = "H",
    integrity: str = "H",
    availability: str = "H",
) -> dict[str, Any]:
    """Calculate CVSS v3.1 base score from metric values.

    Returns a dict with ``score``, ``severity``, ``vector_string``,
    and individual metric scores.
    """
    # Normalise inputs
    av = attack_vector.upper()[:1]
    ac = attack_complexity.upper()[:1]
    pr = privileges_required.upper()[:1]
    ui = user_interaction.upper()[:1]
    s = scope.upper()[:1]
    c = confidentiality.upper()[:1]
    i = integrity.upper()[:1]
    a = availability.upper()[:1]

    for key, val, valid in [
        ("attack_vector", av, _CVSS_METRICS["attack_vector"]),
        ("attack_complexity", ac, _CVSS_METRICS["attack_complexity"]),
        ("privileges_required", pr, _CVSS_METRICS["privileges_required"]),
        ("user_interaction", ui, _CVSS_METRICS["user_interaction"]),
        ("scope", s, _CVSS_METRICS["scope"]),
        ("confidentiality", c, _CVSS_METRICS["confidentiality"]),
        ("integrity", i, _CVSS_METRICS["integrity"]),
        ("availability", a, _CVSS_METRICS["availability"]),
    ]:
        if val not in valid:
            raise ValueError(f"Invalid {key}: {val!r}. Valid: {', '.join(valid)}")

    av_score = _CVSS_METRICS["attack_vector"][av]
    ac_score = _CVSS_METRICS["attack_complexity"][ac]
    pr_score = _CVSS_METRICS["privileges_required"][pr]
    ui_score = _CVSS_METRICS["user_interaction"][ui]
    c_score = _CVSS_METRICS["confidentiality"][c]
    i_score = _CVSS_METRICS["integrity"][i]
    a_score = _CVSS_METRICS["availability"][a]

    # Impact sub-score (ISS)
    iss = 1.0 - (1.0 - c_score) * (1.0 - i_score) * (1.0 - a_score)

    # Impact score
    if s == "U":
        impact = 6.42 * iss
    else:
        impact = 7.52 * (iss - 0.029) - 3.25 * (iss - 0.02) ** 15

    # Exploitability score
    exploitability = 8.22 * av_score * ac_score * pr_score * ui_score

    # Base score
    if impact <= 0:
        base_score = 0.0
    elif s == "U":
        base_score = round(min(impact + exploitability, 10.0), 1)
    else:
        base_score = round(min(1.08 * (impact + exploitability), 10.0), 1)

    # Severity rating
    if base_score >= 9.0:
        severity = "CRITICAL"
    elif base_score >= 7.0:
        severity = "HIGH"
    elif base_score >= 4.0:
        severity = "MEDIUM"
    elif base_score >= 0.1:
        severity = "LOW"
    else:
        severity = "NONE"

    vector = f"CVSS:3.1/AV:{av}/AC:{ac}/PR:{pr}/UI:{ui}/S:{s}/C:{c}/I:{i}/A:{a}"

    return {
        "score": base_score,
        "severity": severity,
        "vector_string": vector,
        "metrics": {
            "attack_vector": {"value": av, "label": _CVSS_LABELS["attack_vector"].get(av, av)},
            "attack_complexity": {"value": ac, "label": _CVSS_LABELS["attack_complexity"].get(ac, ac)},
            "privileges_required": {"value": pr, "label": _CVSS_LABELS["privileges_required"].get(pr, pr)},
            "user_interaction": {"value": ui, "label": _CVSS_LABELS["user_interaction"].get(ui, ui)},
            "scope": {"value": s, "label": _CVSS_LABELS["scope"].get(s, s)},
            "confidentiality": {"value": c, "label": _CVSS_LABELS["confidentiality"].get(c, c)},
            "integrity": {"value": i, "label": _CVSS_LABELS["integrity"].get(i, i)},
            "availability": {"value": a, "label": _CVSS_LABELS["availability"].get(a, a)},
        },
        "scores": {
            "impact": round(impact, 2),
            "exploitability": round(exploitability, 2),
            "iss": round(iss, 4),
        },
    }


def enrich_vulns_with_cvss(vuln_list: list[dict]) -> list[dict]:
    """Add CVSS scores to vulnerability entries that only have severity strings."""
    enriched = []
    for vuln in vuln_list:
        if vuln.get("cvss_score") is not None:
            # Already has a score – use it
            enriched.append(vuln)
            continue

        sev = (vuln.get("cvss_severity", "") or "").upper()
        # Map severity to a representative CVSS vector for scoring
        vec_map = {
            "CRITICAL": {"av": "N", "ac": "L", "pr": "N", "ui": "N", "c": "H", "i": "H", "a": "H"},
            "HIGH":    {"av": "N", "ac": "L", "pr": "L", "ui": "N", "c": "H", "i": "L", "a": "L"},
            "MEDIUM":  {"av": "N", "ac": "L", "pr": "L", "ui": "N", "c": "L", "i": "N", "a": "N"},
            "LOW":     {"av": "A", "ac": "H", "pr": "H", "ui": "R", "c": "L", "i": "N", "a": "N"},
        }
        params = vec_map.get(sev, {"av": "N", "ac": "L", "pr": "N", "ui": "N", "c": "N", "i": "N", "a": "N"})
        try:
            result = calculate_cvss_score(
                attack_vector=params["av"],
                attack_complexity=params["ac"],
                privileges_required=params["pr"],
                user_interaction=params["ui"],
                confidentiality=params["c"],
                integrity=params["i"],
                availability=params["a"],
            )
            vuln["cvss_score"] = result["score"]
            vuln["cvss_vector"] = result["vector_string"]
            vuln["cvss_metrics"] = result["metrics"]
        except (ValueError, KeyError):
            vuln["cvss_score"] = 0.0
            vuln["cvss_vector"] = ""
        enriched.append(vuln)
    return enriched


# ────────────────────────────────────────────────────────────────────────────
# Executive Summary
# ────────────────────────────────────────────────────────────────────────────


def generate_executive_summary(report: dict) -> str:
    """Generate a concise executive summary from the full scan report."""
    target = report.get("target", {})
    scan_info = report.get("scan_info", {})
    sub_info = report.get("subdomain_enumeration", {})
    port_info = report.get("port_scanning", {})
    http_info = report.get("http_probing", {})
    vuln_info = report.get("vulnerability_scan", {})
    ssl_info = report.get("ssl_tls_audit", {})
    takeover_info = report.get("subdomain_takeover", {})
    waf_info = report.get("waf_detection", {})
    loot_info = report.get("loot", {})
    osint_info = report.get("osint", {})

    domain = target.get("domain", "unknown")
    duration = scan_info.get("duration_seconds", 0)
    start_time = scan_info.get("start_time", "")

    total_subdomains = sub_info.get("total_found", 0)
    total_open_ports = sum(h.get("open_ports", 0) for h in port_info.get("hosts", []))
    total_alive = http_info.get("total_alive_services", 0)
    total_cves = vuln_info.get("total_cves", 0)
    total_high_cves = vuln_info.get("total_high_severity", 0)
    total_ssl_issues = sum(
        e.get("total_issues", 0) for e in ssl_info.get("results", {}).values()
    )
    total_takeovers = takeover_info.get("total_vulnerable", 0)
    total_waf = waf_info.get("total_protected", 0)
    total_loot_critical = loot_info.get("critical_count", 0)
    total_osint = osint_info.get("total_findings", 0)
    total_osint_critical = osint_info.get("severity_counts", {}).get("critical", 0)

    lines: list[str] = [
        "=" * 60,
        f"  RECONPROBE EXECUTIVE SUMMARY — {domain}",
        "=" * 60,
        "",
        f"  Scan started:   {start_time or 'N/A'}",
        f"  Duration:       {duration:.1f}s",
        "",
        "  --- Key Findings ---",
        "",
    ]

    # Build findings list
    findings: list[tuple[str, str, int]] = []
    if total_subdomains:
        findings.append(("Subdomains discovered", "info", total_subdomains))
    if total_open_ports:
        findings.append(("Open ports found", "info", total_open_ports))
    if total_alive:
        findings.append(("Live HTTP services", "info", total_alive))
    if total_high_cves:
        findings.append(("High/Critical CVEs", "critical", total_high_cves))
    if total_cves:
        findings.append(("Total CVEs matched", "medium", total_cves))
    if total_ssl_issues:
        findings.append(("SSL/TLS issues", "medium", total_ssl_issues))
    if total_takeovers:
        findings.append(("Subdomain takeovers", "critical", total_takeovers))
    if total_waf:
        findings.append(("WAFs detected", "info", total_waf))
    if total_loot_critical:
        findings.append(("Critical loot items", "critical", total_loot_critical))
    if total_osint_critical:
        findings.append(("Critical OSINT findings", "critical", total_osint_critical))
    if total_osint:
        findings.append(("Total OSINT findings", "info", total_osint))

    if not findings:
        lines.append("  No significant findings recorded.")
    else:
        lines.append("  {:.<45} {:>8}".format("  Finding", "Count"))
        lines.append("  " + "-" * 55)
        for label, severity, count in findings:
            marker = {"critical": "🔴", "high": "🟠", "medium": "🟡", "info": "🔵"}.get(
                severity, "⚪"
            )
            lines.append(f"  {marker} {label:.<42} {count:>8}")

    # Risk assessment
    critical_count = sum(
        c for _, s, c in findings if s == "critical"
    )
    high_count = sum(c for _, s, c in findings if s == "high")
    medium_count = sum(c for _, s, c in findings if s == "medium")

    lines.extend(
        [
            "",
            "  --- Risk Assessment ---",
            "",
            f"  Critical findings:    {critical_count}",
            f"  High findings:        {high_count}",
            f"  Medium findings:      {medium_count}",
            "",
            "  --- Recommendations ---",
            "",
        ]
    )

    recs: list[str] = []
    if total_takeovers:
        recs.append(
            "  🔴 Immediately investigate and remediate subdomain takeover "
            "vulnerabilities — these can lead to full domain compromise."
        )
    if total_loot_critical:
        recs.append(
            "  🔴 Review and rotate any exposed credentials, API keys, or "
            "tokens found in loot collection."
        )
    if total_osint_critical:
        recs.append(
            "  🔴 Address critical OSINT findings — sensitive information "
            "exposed through public sources."
        )
    if total_high_cves:
        recs.append(
            f"  🟠 Patch {total_high_cves} high/critical severity "
            "vulnerabilities as soon as possible."
        )
    if total_ssl_issues:
        recs.append(
            f"  🟠 Review {total_ssl_issues} SSL/TLS issues — disable weak "
            "protocols and ciphers."
        )
    if total_open_ports > 50:
        recs.append(
            "  🟡 Reduce attack surface by closing unnecessary open ports."
        )

    if not recs:
        recs.append("  ✅ No critical recommendations at this time.")

    lines.extend(recs)
    lines.extend(
        [
            "",
            "=" * 60,
            f"  Report generated: {datetime.now(timezone.utc).isoformat()}",
            f"  Tool: ReconProbe v{report.get('version', '')}",
            "=" * 60,
        ]
    )

    return "\n".join(lines)


# ────────────────────────────────────────────────────────────────────────────
# Scan Timeline
# ────────────────────────────────────────────────────────────────────────────


def generate_timeline_entries(report: dict) -> list[dict[str, Any]]:
    """Build a list of timeline events from the scan report for chronological display."""
    entries: list[dict[str, Any]] = []
    scan_info = report.get("scan_info", {})

    start = scan_info.get("start_time", "")

    # Determine which phases had data
    phase_markers: list[tuple[str, str, str]] = [
        ("subdomain_enumeration", "Subdomain Enumeration", "subdomains"),
        ("advanced_subdomain_techniques", "Advanced Subdomain Techniques", "advanced"),
        ("port_scanning", "Port Scanning", "ports"),
        ("http_probing", "HTTP Probing & Fingerprinting", "HTTP"),
        ("enrichment", "Enrichment (Shodan/NVD)", "enrichment"),
        ("crawling", "Web Crawling", "crawling"),
        ("dirbuster", "Directory Brute-Force", "dirbuster"),
        ("screenshots", "Screenshot Capture", "screenshots"),
        ("vulnerability_scan", "Vulnerability Scan", "vulns"),
        ("ssl_tls_audit", "SSL/TLS Audit", "SSL"),
        ("subdomain_takeover", "Subdomain Takeover Detection", "takeover"),
        ("waf_detection", "WAF Detection", "WAF"),
        ("exploit_suggestions", "Exploit Suggestion", "exploits"),
        ("payloads", "Payload Generation", "payloads"),
        ("loot", "Loot Collection", "loot"),
        ("msf_script", "MSF Script Generation", "MSF"),
        ("osint", "Advanced OSINT", "OSINT"),
    ]

    for key, label, _ in phase_markers:
        data = report.get(key, {})
        if data and any(
            isinstance(v, (dict, list)) and v
            for v in data.values()
            if not isinstance(v, str)
        ):
            entries.append(
                {
                    "phase": label,
                    "timestamp": start if start else "",
                    "status": "completed",
                }
            )

    return entries


# ────────────────────────────────────────────────────────────────────────────
# CSV Export
# ────────────────────────────────────────────────────────────────────────────


def export_findings_csv(report: dict) -> str:
    """Export all findings from the scan report as a CSV string."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    writer.writerow(
        [
            "Category",
            "Type",
            "Severity",
            "Target",
            "Value",
            "Description",
            "CVSS Score",
            "CVE ID",
        ]
    )

    # Subdomains
    for r in report.get("subdomain_enumeration", {}).get("results", []):
        writer.writerow(
            [
                "Subdomain",
                "subdomain",
                "info",
                r.get("hostname", ""),
                r.get("ip_address", ""),
                f"Source: {r.get('source', '')}",
                "",
                "",
            ]
        )

    # Ports
    for host in report.get("port_scanning", {}).get("hosts", []):
        for p in host.get("ports", []):
            if p.get("state") == "open":
                writer.writerow(
                    [
                        "Port",
                        "open_port",
                        "info",
                        host.get("hostname", ""),
                        f"{p.get('port', '')}/{p.get('service', '')}",
                        p.get("banner", "") or "",
                        "",
                        "",
                    ]
                )

    # CVEs from vulnerability scan
    for cve in report.get("vulnerability_scan", {}).get("cve_matches", []):
        writer.writerow(
            [
                "Vulnerability",
                "cve_match",
                cve.get("cvss_severity", "NONE"),
                cve.get("affected_service", ""),
                cve.get("cve_id", ""),
                (cve.get("description", "") or "")[:200],
                cve.get("cvss_score", ""),
                cve.get("cve_id", ""),
            ]
        )

    # Default creds
    for cred in report.get("vulnerability_scan", {}).get("default_credentials", []):
        writer.writerow(
            [
                "Credential",
                "default_cred",
                "high",
                cred.get("hostname", ""),
                f"{cred.get('username', '')}:{cred.get('password', '')}",
                f"Default creds for {cred.get('service', '')}",
                "",
                "",
            ]
        )

    # SSL issues
    for endpoint_key, endpoint_data in (
        report.get("ssl_tls_audit", {}).get("results", {}).items()
    ):
        for wc in endpoint_data.get("weak_ciphers", []):
            writer.writerow(
                [
                    "SSL/TLS",
                    "weak_cipher",
                    "medium",
                    endpoint_key,
                    wc.get("cipher", ""),
                    wc.get("reason", ""),
                    "",
                    "",
                ]
            )
        for proto in endpoint_data.get("protocols", []):
            if proto.get("protocol") in ("TLS 1.0", "TLS 1.1") and proto.get(
                "supported"
            ):
                writer.writerow(
                    [
                        "SSL/TLS",
                        "outdated_protocol",
                        "medium",
                        endpoint_key,
                        proto.get("protocol", ""),
                        "Outdated TLS protocol supported",
                        "",
                        "",
                    ]
                )

    # Takeover results
    for r in report.get("subdomain_takeover", {}).get("results", []):
        writer.writerow(
            [
                "Takeover",
                "subdomain_takeover",
                "critical",
                r.get("hostname", ""),
                r.get("service", ""),
                f"Confidence: {r.get('confidence', '')}",
                "",
                "",
            ]
        )

    # WAF results
    for url_str, result in report.get("waf_detection", {}).get("results", {}).items():
        if result.get("is_protected"):
            waf_names = [w["name"] for w in result.get("detected_wafs", [])]
            writer.writerow(
                [
                    "WAF",
                    "waf_detected",
                    "info",
                    url_str,
                    ", ".join(waf_names),
                    "WAF fingerprinting completed",
                    "",
                    "",
                ]
            )

    # Exploit suggestions
    for s in report.get("exploit_suggestions", {}).get("suggestions", []):
        writer.writerow(
            [
                "Exploit",
                "exploit_suggestion",
                s.get("reliability", "medium"),
                s.get("service", ""),
                s.get("edb_id", ""),
                s.get("title", "")[:200],
                "",
                s.get("cve_id", ""),
            ]
        )

    # Loot items
    for item in report.get("loot", {}).get("items", []):
        writer.writerow(
            [
                "Loot",
                item.get("type", ""),
                item.get("severity", "info"),
                item.get("target", ""),
                str(item.get("data", ""))[:100],
                item.get("description", ""),
                "",
                "",
            ]
        )

    # OSINT findings
    for finding in report.get("osint", {}).get("findings", []):
        writer.writerow(
            [
                "OSINT",
                finding.get("type", ""),
                finding.get("severity", "info"),
                finding.get("source", ""),
                (finding.get("value", "") or "")[:100],
                (finding.get("context", "") or "")[:200],
                "",
                "",
            ]
        )

    return output.getvalue()


def export_findings_xlsx(report: dict, output_path: Path) -> None:
    """Export all findings to an XLSX workbook."""
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError(
            "openpyxl is required for XLSX export. Install it with: "
            "pip install openpyxl"
        )

    wb = Workbook()
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1d29", end_color="1a1d29", fill_type="solid")
    critical_fill = PatternFill(start_color="450a0a", end_color="450a0a", fill_type="solid")
    high_fill = PatternFill(start_color="7f1d1d", end_color="7f1d1d", fill_type="solid")
    medium_fill = PatternFill(start_color="713f12", end_color="713f12", fill_type="solid")

    def _style_header(ws, headers: list[str]) -> None:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # ── Summary Sheet ──
    ws_summary = wb.active
    assert ws_summary is not None, "Workbook must have an active sheet"
    ws_summary.title = "Summary"
    summary_data = [
        ("Target Domain", report.get("target", {}).get("domain", "")),
        ("Scan Start", report.get("scan_info", {}).get("start_time", "")),
        ("Duration (s)", report.get("scan_info", {}).get("duration_seconds", 0)),
        ("Subdomains Found", report.get("subdomain_enumeration", {}).get("total_found", 0)),
        ("Open Ports", sum(h.get("open_ports", 0) for h in report.get("port_scanning", {}).get("hosts", []))),
        ("CVEs Found", report.get("vulnerability_scan", {}).get("total_cves", 0)),
        ("Takeover Vulns", report.get("subdomain_takeover", {}).get("total_vulnerable", 0)),
        ("Loot Items", report.get("loot", {}).get("total_count", 0)),
        ("OSINT Findings", report.get("osint", {}).get("total_findings", 0)),
    ]
    for row_idx, (key, val) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2, value=val)
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 40

    # ── CVEs Sheet ──
    cves = report.get("vulnerability_scan", {}).get("cve_matches", [])
    if cves:
        ws_cves = wb.create_sheet("CVEs")
        headers = ["CVE ID", "Severity", "CVSS", "Service", "Description"]
        _style_header(ws_cves, headers)
        for row_idx, cve in enumerate(cves, 2):
            ws_cves.cell(row=row_idx, column=1, value=cve.get("cve_id", ""))
            sev_cell = ws_cves.cell(
                row=row_idx, column=2, value=cve.get("cvss_severity", "")
            )
            sev = (cve.get("cvss_severity", "") or "").upper()
            if sev == "CRITICAL":
                sev_cell.fill = critical_fill
            elif sev == "HIGH":
                sev_cell.fill = high_fill
            elif sev == "MEDIUM":
                sev_cell.fill = medium_fill
            score = cve.get("cvss_score")
            ws_cves.cell(row=row_idx, column=3, value=round(score, 1) if score is not None else "")
            ws_cves.cell(row=row_idx, column=4, value=cve.get("affected_service", ""))
            ws_cves.cell(
                row=row_idx, column=5, value=(cve.get("description", "") or "")[:200]
            )
        for col_idx, h in enumerate(headers, 1):
            ws_cves.column_dimensions[chr(64 + col_idx)].width = max(
                15, len(h) + 5
            )

    # ── Loot Sheet ──
    loot_items = report.get("loot", {}).get("items", [])
    if loot_items:
        ws_loot = wb.create_sheet("Loot")
        headers = ["Type", "Source", "Target", "Severity", "Description", "Data"]
        _style_header(ws_loot, headers)
        for row_idx, item in enumerate(loot_items, 2):
            ws_loot.cell(row=row_idx, column=1, value=item.get("type", ""))
            ws_loot.cell(row=row_idx, column=2, value=item.get("source", ""))
            ws_loot.cell(row=row_idx, column=3, value=item.get("target", ""))
            sev = item.get("severity", "info")
            sev_cell = ws_loot.cell(row=row_idx, column=4, value=sev)
            if sev == "critical":
                sev_cell.fill = critical_fill
            elif sev == "high":
                sev_cell.fill = high_fill
            ws_loot.cell(row=row_idx, column=5, value=(item.get("description", "") or "")[:200])
            ws_loot.cell(row=row_idx, column=6, value=str(item.get("data", ""))[:100])
        for col_idx in range(1, 7):
            ws_loot.column_dimensions[chr(64 + col_idx)].width = 20

    # ── OSINT Sheet ──
    osint_findings = report.get("osint", {}).get("findings", [])
    if osint_findings:
        ws_osint = wb.create_sheet("OSINT")
        headers = ["Source", "Type", "Severity", "Value", "Context"]
        _style_header(ws_osint, headers)
        for row_idx, f in enumerate(osint_findings, 2):
            ws_osint.cell(row=row_idx, column=1, value=f.get("source", ""))
            ws_osint.cell(row=row_idx, column=2, value=f.get("type", ""))
            sev_cell = ws_osint.cell(
                row=row_idx, column=3, value=f.get("severity", "")
            )
            sev = (f.get("severity", "") or "").upper()
            if sev == "CRITICAL":
                sev_cell.fill = critical_fill
            elif sev == "HIGH":
                sev_cell.fill = high_fill
            ws_osint.cell(
                row=row_idx, column=4, value=(f.get("value", "") or "")[:120]
            )
            ws_osint.cell(
                row=row_idx, column=5, value=(f.get("context", "") or "")[:200]
            )
        for col_idx in range(1, 6):
            ws_osint.column_dimensions[chr(64 + col_idx)].width = 22

    # ── SSL/TLS Sheet ──
    ssl_results = report.get("ssl_tls_audit", {}).get("results", {})
    ssl_issue_rows: list[tuple[str, str, str]] = []
    for ep, ed in ssl_results.items():
        for wc in ed.get("weak_ciphers", []):
            ssl_issue_rows.append((ep, wc.get("cipher", ""), wc.get("reason", "")))
        for proto in ed.get("protocols", []):
            if proto.get("protocol") in ("TLS 1.0", "TLS 1.1") and proto.get("supported"):
                ssl_issue_rows.append(
                    (ep, proto.get("protocol", ""), "Outdated protocol supported")
                )
    if ssl_issue_rows:
        ws_ssl = wb.create_sheet("SSL Issues")
        headers = ["Endpoint", "Issue", "Details"]
        _style_header(ws_ssl, headers)
        for row_idx, (ep, issue, detail) in enumerate(ssl_issue_rows, 2):
            ws_ssl.cell(row=row_idx, column=1, value=ep)
            ws_ssl.cell(row=row_idx, column=2, value=issue)
            ws_ssl.cell(row=row_idx, column=3, value=detail)
        ws_ssl.column_dimensions["A"].width = 35
        ws_ssl.column_dimensions["B"].width = 30
        ws_ssl.column_dimensions["C"].width = 30

    wb.save(str(output_path))


# ────────────────────────────────────────────────────────────────────────────
# PDF Report Generation
# ────────────────────────────────────────────────────────────────────────────


if FPDF_AVAILABLE:

    class _ReconPDF(FPDF):
        """Custom FPDF subclass with headers/footers for ReconProbe reports."""

        def header(self) -> None:
            self.set_font("Helvetica", "B", 10)
            self.set_text_color(6, 182, 212)
            self.cell(0, 8, "ReconProbe — Security Assessment Report", new_x="LMARGIN", new_y="NEXT", align="C")
            self.line(10, self.get_y(), 200, self.get_y())
            self.ln(4)

        def footer(self) -> None:
            self.set_y(-15)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(100, 116, 139)
            self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")

        def section_title(self, title: str) -> None:
            self.set_font("Helvetica", "B", 14)
            self.set_text_color(15, 23, 42)
            self.set_fill_color(226, 232, 240)
            self.cell(0, 10, f"  {title}", new_x="LMARGIN", new_y="NEXT", fill=True)
            self.ln(4)

        def severity_badge(self, severity: str) -> str:
            sev = severity.upper()
            if sev == "CRITICAL":
                return "[ CRITICAL ]"
            elif sev == "HIGH":
                return "[ HIGH ]"
            elif sev == "MEDIUM":
                return "[ MEDIUM ]"
            elif sev == "LOW":
                return "[ LOW ]"
            return "[ INFO ]"

        def write_finding(
            self, label: str, value: str, severity: str = ""
        ) -> None:
            self.set_font("Helvetica", "B", 10)
            self.set_text_color(55, 65, 81)
            badge = self.severity_badge(severity) if severity else ""
            self.cell(0, 6, f"{badge} {label}", new_x="LMARGIN", new_y="NEXT")
            self.set_font("Helvetica", "", 9)
            self.set_text_color(107, 114, 128)
            self.multi_cell(0, 5, value)
            self.ln(2)

        def write_table(
            self, headers: list[str], rows: list[list[str]], col_widths: Optional[list[int]] = None
        ) -> None:
            if not rows:
                return
            if col_widths is None:
                col_widths = [int(180 / len(headers))] * len(headers)

            # Header row
            self.set_font("Helvetica", "B", 8)
            self.set_fill_color(15, 23, 42)
            self.set_text_color(255, 255, 255)
            for i, header in enumerate(headers):
                w = col_widths[i] if i < len(col_widths) else 40
                self.cell(w, 7, header, border=1, fill=True, align="C")
            self.ln()

            # Data rows
            self.set_font("Helvetica", "", 7)
            self.set_text_color(55, 65, 81)
            for row in rows:
                for i, cell_val in enumerate(row):
                    w = col_widths[i] if i < len(col_widths) else 40
                    self.cell(w, 5, str(cell_val), border=1)
                self.ln()
            self.ln(3)


def generate_pdf_report(report: dict, output_path: Path) -> Path:
    """Generate a professional PDF report from the scan report dict."""
    if not FPDF_AVAILABLE:
        raise RuntimeError(
            "fpdf2 is required for PDF generation. Install it with: pip install fpdf2"
        )

    pdf = _ReconPDF()
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    target = report.get("target", {})
    scan_info = report.get("scan_info", {})
    domain = target.get("domain", "unknown")
    duration = scan_info.get("duration_seconds", 0)
    start_time = scan_info.get("start_time", "")

    # ── Title Page ──
    pdf.ln(40)
    pdf.set_font("Helvetica", "B", 28)
    pdf.set_text_color(6, 182, 212)
    pdf.cell(0, 15, "Security Assessment Report", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)
    pdf.set_font("Helvetica", "", 16)
    pdf.set_text_color(100, 116, 139)
    pdf.cell(0, 10, f"Target: {domain}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 8, f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Duration: {duration:.1f}s  |  Start: {start_time[:19] if start_time else 'N/A'}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Tool: ReconProbe v{report.get('version', '')}", align="C", new_x="LMARGIN", new_y="NEXT")

    # ── Executive Summary ──
    pdf.add_page()
    pdf.section_title("Executive Summary")
    summary_text = generate_executive_summary(report)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(55, 65, 81)
    for line in summary_text.split("\n"):
        line = line.strip()
        if not line:
            pdf.ln(2)
            continue
        # Strip markdown/emoji/box-drawing chars for PDF
        clean = (
            line.replace("=" * 60, "")
            .replace("🔴", "[!]")
            .replace("🟠", "[!]")
            .replace("🟡", "[~]")
            .replace("🔵", "[i]")
            .replace("⚪", "[*]")
        )
        if clean:
            pdf.multi_cell(0, 4.5, clean)
    pdf.ln(4)

    # ── Subdomains ──
    sub_info = report.get("subdomain_enumeration", {})
    results = sub_info.get("results", [])
    if results:
        pdf.add_page()
        pdf.section_title(f"Subdomain Enumeration ({sub_info.get('total_found', 0)} found)")
        headers = ["Hostname", "IP", "Source"]
        rows = [
            [r.get("hostname", "")[:50], r.get("ip_address", "") or "-", r.get("source", "")]
            for r in results[:60]
        ]
        pdf.write_table(headers, rows, [70, 40, 40])
        if len(results) > 60:
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_text_color(100, 116, 139)
            pdf.cell(0, 5, f"... and {len(results) - 60} more subdomains", new_x="LMARGIN", new_y="NEXT")

    # ── Ports ──
    port_info = report.get("port_scanning", {})
    hosts = port_info.get("hosts", [])
    total_open = sum(h.get("open_ports", 0) for h in hosts)
    if total_open:
        pdf.add_page()
        pdf.section_title(f"Port Scanning ({total_open} open ports)")
        for host in hosts:
            open_ports = [p for p in host.get("ports", []) if p.get("state") == "open"]
            if not open_ports:
                continue
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(6, 182, 212)
            pdf.cell(0, 6, f"{host.get('hostname', '')} ({host.get('ip_address', '')})", new_x="LMARGIN", new_y="NEXT")
            headers = ["Port", "Service", "Banner"]
            rows = [
                [
                    str(p.get("port", "")),
                    p.get("service", ""),
                    (p.get("banner", "") or "")[:40],
                ]
                for p in open_ports[:30]
            ]
            pdf.write_table(headers, rows, [20, 30, 80])

    # ── CVEs (Vulnerability Scan) ──
    vuln_info = report.get("vulnerability_scan", {})
    cves = vuln_info.get("cve_matches", [])
    if cves:
        pdf.add_page()
        pdf.section_title(f"Vulnerability Scan — CVE Matches ({len(cves)})")
        headers = ["CVE ID", "Severity", "CVSS", "Service", "Description"]
        rows = []
        for cve in cves[:40]:
            score = cve.get("cvss_score")
            score_str = f"{score:.1f}" if score is not None else "-"
            rows.append(
                [
                    cve.get("cve_id", ""),
                    (cve.get("cvss_severity", "") or "")[:8],
                    score_str,
                    cve.get("affected_service", ""),
                    (cve.get("description", "") or "")[:60],
                ]
            )
        pdf.write_table(headers, rows, [30, 20, 15, 25, 60])

    # ── SSL/TLS ──
    ssl_info = report.get("ssl_tls_audit", {})
    ssl_results = ssl_info.get("results", {})
    if ssl_results:
        pdf.add_page()
        pdf.section_title("SSL/TLS Audit")
        for ep, ed in ssl_results.items():
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(6, 182, 212)
            grade = ed.get("grade", "?")
            pdf.cell(0, 6, f"{ep}  —  Grade: {grade}", new_x="LMARGIN", new_y="NEXT")

            cert = ed.get("certificate", {})
            if cert:
                pdf.set_font("Helvetica", "", 9)
                pdf.set_text_color(55, 65, 81)
                pdf.cell(0, 5, f"  Subject: {cert.get('subject', '')}", new_x="LMARGIN", new_y="NEXT")
                pdf.cell(0, 5, f"  Issuer:  {cert.get('issuer', '')}", new_x="LMARGIN", new_y="NEXT")
                pdf.cell(0, 5, f"  Valid To: {cert.get('valid_to', '')}", new_x="LMARGIN", new_y="NEXT")
                if cert.get("is_expired"):
                    pdf.set_text_color(239, 68, 68)
                    pdf.cell(0, 5, "  ** EXPIRED **", new_x="LMARGIN", new_y="NEXT")
                    pdf.set_text_color(55, 65, 81)

            wc = ed.get("weak_ciphers", [])
            if wc:
                pdf.write_table(
                    ["Cipher", "Reason"],
                    [[c.get("cipher", ""), c.get("reason", "")] for c in wc[:10]],
                    [60, 80],
                )
            pdf.ln(2)

    # ── Takeover ──
    takeover_info = report.get("subdomain_takeover", {})
    if takeover_info.get("total_vulnerable", 0) > 0:
        pdf.add_page()
        pdf.section_title("Subdomain Takeover Detection")
        headers = ["Hostname", "Service", "Confidence"]
        rows = [
            [
                r.get("hostname", ""),
                r.get("service", ""),
                r.get("confidence", ""),
            ]
            for r in takeover_info.get("results", [])
        ]
        pdf.write_table(headers, rows, [70, 40, 30])

    # ── Loot ──
    loot_info = report.get("loot", {})
    critical_items = [i for i in loot_info.get("items", []) if i.get("severity") == "critical"]
    if critical_items:
        pdf.add_page()
        pdf.section_title(f"Critical Loot Items ({len(critical_items)})")
        for item in critical_items[:15]:
            pdf.write_finding(
                f"{item.get('type', '')} — {item.get('target', '')}",
                str(item.get("data", ""))[:150],
                "critical",
            )

    # ── OSINT ──
    osint_info = report.get("osint", {})
    osint_findings = osint_info.get("findings", [])
    if osint_findings:
        pdf.add_page()
        pdf.section_title(f"OSINT Findings ({len(osint_findings)})")
        for finding in osint_findings[:30]:
            sev = finding.get("severity", "info")
            pdf.write_finding(
                f"[{finding.get('source', '')}] {finding.get('type', '')}",
                str(finding.get("value", "") or "")[:120],
                sev,
            )

    # ── Export ──
    pdf.output(str(output_path))
    return output_path


# ────────────────────────────────────────────────────────────────────────────
# Convenience: run all Phase 6 enhancements on a report
# ────────────────────────────────────────────────────────────────────────────


def run_reporting_enhancements(
    report: dict,
    output_dir: Path,
    generate_pdf: bool = False,
    generate_csv: bool = False,
    generate_xlsx: bool = False,
    generate_exec_summary: bool = False,
) -> dict[str, Path]:
    """Run Phase 6 reporting enhancements on a completed scan report.

    Returns a dict mapping report type to paths generated.
    """
    generated: dict[str, Path] = {}
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    domain_safe = report.get("target", {}).get("domain", "unknown").replace(".", "_")

    # Enrich CVEs with CVSS scores
    vuln_info = report.get("vulnerability_scan", {})
    if vuln_info.get("cve_matches"):
        vuln_info["cve_matches"] = enrich_vulns_with_cvss(vuln_info["cve_matches"])

    if generate_pdf:
        pdf_path = output_dir / f"{domain_safe}_report.pdf"
        try:
            generate_pdf_report(report, pdf_path)
            generated["pdf"] = pdf_path
        except RuntimeError:
            # fpdf2 not available
            pass

    if generate_csv:
        csv_path = output_dir / f"{domain_safe}_findings.csv"
        csv_content = export_findings_csv(report)
        csv_path.write_text(csv_content)
        generated["csv"] = csv_path

    if generate_xlsx:
        xlsx_path = output_dir / f"{domain_safe}_findings.xlsx"
        try:
            export_findings_xlsx(report, xlsx_path)
            generated["xlsx"] = xlsx_path
        except RuntimeError:
            pass

    if generate_exec_summary:
        summary_path = output_dir / f"{domain_safe}_executive_summary.txt"
        summary = generate_executive_summary(report)
        summary_path.write_text(summary)
        generated["exec_summary"] = summary_path

    return generated
